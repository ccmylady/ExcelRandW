# 未完
# 文件打开错误保护
# 信息写入前比较保护
# 序号唯一性检索
# 格式保护
# 默认输入

import logging
logging.basicConfig(level=logging.INFO,format='%(asctime)s - %(levelname)s - %(message)s')
logging.info('程序开始')
logging.info('数据汇集工具，第1版，20220622。\n'
             '将分所填充后的表格某些项汇总填入总表,打开1个文件保存1次模式\n'
             '注意事项：1）所有表格前内容必须一致\n'
             '2）序号必须是唯一的\n'
             '3) 数据源文件必须是xls格式，转化后写入文件必须是xlsx格式\n'
             '4) 表名固定 \n'
             '5) 文件存放于C:\EXCELzz\origin \n'
             '6) 汇总后文件存放于C:\EXCELzz\summary \n')

import openpyxl
import xlrd
import os

def main():
    #数据源文件，及汇总文件夹地址
    files_orgin_path=r'C:\EXCELzz\origin'
    files_summary_path=r'C:\EXCELzz\summary'

    #多次循环输入所名和文件名
    files={}
    while True:
        files_input=input('\n请输入所名和文件名，以‘xxxx所：文件名’的形式，或输入over,以结束输入:\n')
        if files_input=="over":
            break
        office_input=files_input.split(":",1)
        files[office_input[0]]=os.path.join(files_orgin_path,office_input[1])

    # 国家税务总局上海市松江区税务局第十一税务所:附表1、一般纳税人小微企业印花税申报应享未享“六税两费”减免政策（需反馈）(11所).xls
    # files={'国家税务总局上海市松江区税务局第十一税务所':'C:\EXCELzz\origin\附表1、一般纳税人小微企业印花税申报应享未享“六税两费”减免政策（需反馈）(11所).xls',
    #        '国家税务总局上海市松江区税务局第十五税务所':'C:\EXCELzz\origin\附表1、一般纳税人小微企业印花税申报应享未享“六税两费”减免政策（需反馈）(十五所).xls'
    #        }

    logging.info(files)

    #输入需要汇总的表格字段，当前以列号
    keywords_input = input('\n请输入汇总信息所在列序号，以‘列号1,列号2,列号3,...’的形式(如默认36,37,38请直接按回车):\n')
    if len(keywords_input)==0:
        keywords_input='36,37,38'
    keywords = keywords_input.split(",")
    keywords = [int(x) for x in keywords]
    logging.info("汇总信息所在列序号:{}".format(keywords))

    #请输入数据项唯一标识符(ID)所在列序号
    index_ID_input=input('\n请输入索引号（ID）所在列序号(如默认1请直接按回车):\n')
    if len(index_ID_input)==0:
        index_ID_input='1'
    index_num=int(index_ID_input)
    logging.info("索引号（ID）所在列序号:{}".format(index_num))

    #请输入税务所名所在列序号
    index_office_input=input('\n请输入税务所名所在列序号(如默认6请直接按回车):\n')
    if len(index_office_input)==0:
        index_office_input='6'
    index_office_num=int(index_office_input)
    logging.info("税务所名所在列序号:{}".format(index_office_num))


    #请输入信息所在表名
    sheetname_input=input("\n请输入信息所在表名(如默认'SQL Results'请直接按回车):\n")
    if len(sheetname_input)==0:
        sheetname_input='SQL Results'
    logging.info("信息所在表名:{}".format(sheetname_input))

    # 请输入汇总表模版名
    files_summarytemplate_name=input("\n请输入汇总表模版名(如默认'summary template.xlsx'请直接按回车):\n")
    if len(files_summarytemplate_name)==0:
        files_summarytemplate_name='summary template.xlsx'
    files_summarytemplate_fullpath = os.path.join(files_summary_path,files_summarytemplate_name)
    logging.info("汇总表模版名:{}".format(files_summarytemplate_fullpath))

    # 请输入汇总表表名
    files_summary_name=input("\n请输入汇总表名(如默认'summary.xlsx'请直接按回车):\n")
    if len(files_summary_name)==0:
        files_summary_name='summary.xlsx'
    files_summary_fullpath = os.path.join(files_summary_path,files_summary_name)
    logging.info("汇总表名:{}".format(files_summary_fullpath))

    contents_update = {}
    num=0
    #对于待汇总的每个所，读取每个所的更新信息
    for office in files.keys():
        # logging.debug(office+':'+files[office])
        # if os.path.splitext(filename_purchase)[1] in ['.XLSX','.xlsx']:
        #打开源文件
        workbook_r = xlrd.open_workbook(filename=files[office])
        worksheet_r = workbook_r.sheet_by_name(sheet_name=sheetname_input)
        rows = worksheet_r.nrows

        #读取表单中的每一行
        for rowNum1 in range(rows):
            worksheet_list = worksheet_r.row_values(rowx=rowNum1, start_colx=0, end_colx=None)
            #如果该行为待汇总所在行，则添加到待更新库中
            if worksheet_list[index_office_num]== office:
                contents_update[worksheet_list[(index_num-1)]]=worksheet_list

        logging.debug(contents_update)
        logging.info("{} 共有 {} 条信息需要更新".format(office,(len(contents_update.keys())-num)))
        num=len(contents_update.keys())

    logging.info("共有 {} 条信息需要更新".format(len(contents_update.keys())))

    workbook_w=openpyxl.load_workbook(files_summarytemplate_fullpath)
    worksheet_w=workbook_w.get_sheet_by_name(sheetname_input)

    # 更新至表中
    updateNum_total=0

    for rowNum2 in range(2, worksheet_w.max_row + 1):
        serialNum=worksheet_w.cell(row=rowNum2,column=index_num).value
        if serialNum in contents_update.keys():
            # 校验机制1
            flag_update = 1
            for j in range(len(contents_update[serialNum])):
                if (j+1) in keywords:
                    continue
                logging.debug(contents_update[serialNum][j])
                logging.debug(worksheet_w.cell(row=rowNum2,column=(j+1)).value)
                if contents_update[serialNum][j]=='' and (worksheet_w.cell(row=rowNum2,column=(j+1)).value is None):
                    continue
                if contents_update[serialNum][j] != worksheet_w.cell(row=rowNum2,column=(j+1)).value:
                    logging.info("序号{}项{}列，原始表与汇总表数据不一致，暂不更新请复核".format(serialNum,j+1))
                    flag_update=0
                    break
            if flag_update:
                for keyword in keywords:
                    worksheet_w.cell(row=rowNum2,column=keyword).value=contents_update[serialNum][(keyword-1)]
                updateNum_total+=1

    # 可以考虑下多次读取后，1次写入
    workbook_w.save(files_summary_fullpath)
    logging.info("共有 {} 条信息已更新".format(updateNum_total))

    logging.info('程序结束')

    input('请输入任意键退出')

if __name__ == '__main__':
    main()


