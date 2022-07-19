# 未完
# 文件打开错误保护
# 信息写入前比较保护
# 序号唯一性检索
# 格式保护

import logging
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')
logging.info('start of program')
logging.info('数据汇集工具，第1版，20220622。\n'
             '将分所填充后的表格某些项汇总填入总表,打开1个文件保存1次模式\n'
             '注意事项：1）所有表格前内容必须一致\n'
             '2）序号必须是唯一的\n'
             '3) 数据源文件必须是xls格式，转化后写入文件必须是xlsx格式\n'
             '4 表名固定 \n' )

import openpyxl
import xlrd
import os

def main():
    #数据源文件，及汇总文件夹地址
    files_orgin_path=r'C:\EXCELzz\origin'
    files_summary_path=r'C:\EXCELzz\summary'

    #多次循环输入所名和文件名
    files={}
    while True:
        files_input=input('\n请输入所名和文件名，以‘xxxx所：文件名’的形式，或输入over,以结束输入:\n')
        if files_input=="over":
            break
        office_input=files_input.split(":",1)
        files[office_input[0]]=os.path.join(files_orgin_path,office_input[1])

    files={'国家税务总局上海市松江区税务局第十一税务所':'C:\EXCELzz\origin\附表1、一般纳税人小微企业印花税申报应享未享“六税两费”减免政策（需反馈）(11所).xls',
           '国家税务总局上海市松江区税务局第十五税务所':'C:\EXCELzz\origin\附表1、一般纳税人小微企业印花税申报应享未享“六税两费”减免政策（需反馈）(十五所).xls'
           }

    logging.debug(files)

    #输入需要汇总的表格字段，当前以列号
    keywords_input = input('\n请输入汇总信息所在列序号，以‘列号1,列号2,列号3,...’的形式:\n')
    keywords = keywords_input.split(",")
    keywords = [int(x) for x in keywords]
    print(keywords)
    logging.debug(keywords)

    #请输入数据项唯一标识符(ID)所在列序号
    index_input=input('\n请输入索引号（ID）所在列编号:\n')
    index_num=int(index_input)

    #对于待汇总的每个源文件
    for office in files.keys():
        # logging.debug(office+':'+files[office])
        # if os.path.splitext(filename_purchase)[1] in ['.XLSX','.xlsx']:
        workbook_r = xlrd.open_workbook(filename=files[office])
        worksheet_r = workbook_r.sheet_by_name(sheet_name='SQL Results')
        rows = worksheet_r.nrows

        contents_update={}
        for rowNum1 in range(rows):
            worksheet_list = worksheet_r.row_values(rowx=rowNum1, start_colx=0, end_colx=None)
            if worksheet_list[6]== office:
                contents_update[worksheet_list[0]]=worksheet_list

        logging.debug(contents_update)
        logging.info("{} 共有 {} 条信息需要更新".format(office,len(contents_update.keys())))

        workbook_w=openpyxl.load_workbook(r'C:\EXCELzz\summary\附表1、一般纳税人小微企业印花税申报应享未享“六税两费”减免政策（需反馈）(汇总表).xlsx')
        worksheet_w=workbook_w.get_sheet_by_name('SQL Results')

        updateNum=0
        for rowNum2 in range(2, worksheet_w.max_row + 1):
            serialNum=worksheet_w.cell(row=rowNum2,column=index_num).value
            if serialNum in contents_update.keys():
                # 数字待改
                for keyword in keywords:
                    worksheet_w.cell(row=rowNum2,column=keyword).value=contents_update[serialNum][(keyword-1)]
                    # worksheet_w.cell(row=rowNum2, column=37).value = contents_update[serialNum][36]
                    # worksheet_w.cell(row=rowNum2, column=38).value = contents_update[serialNum][37]
                updateNum+=1

        # 可以考虑下多次读取后，1次写入
        workbook_w.save(r'C:\EXCELzz\summary\附表1、一般纳税人小微企业印花税申报应享未享“六税两费”减免政策（需反馈）(汇总表).xlsx')
        logging.info("{} 共有 {} 条信息已更新".format(office,updateNum))

    logging.debug('end of program')

if __name__ == '__main__':
    main()


